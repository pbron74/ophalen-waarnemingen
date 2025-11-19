import pandas as pd
import folium
from tkinter import Tk, filedialog, messagebox, Label, Button, StringVar, OptionMenu, Entry
from geopy.distance import geodesic
from datetime import datetime
import math
import openpyxl
import re
import webbrowser
import os

bestand_pad = None  # Globale variabele voor geselecteerd bestand

def parse_gps(gps_str):
    match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', str(gps_str))
    if match:
        return float(match.group(1)), float(match.group(2))
    return None

def genereer_vallen(coord, afstand):
    vallen = []
    for hoek in range(0, 360, 45):
        dx = afstand * math.cos(math.radians(hoek))
        dy = afstand * math.sin(math.radians(hoek))
        lat_offset = dy / 111000
        lon_offset = dx / (111000 * math.cos(math.radians(coord[0])))
        vallen.append((coord[0] + lat_offset, coord[1] + lon_offset))
    return vallen

def lees_excel_met_links(filepath):
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()

    for col in df.columns:
        if str(col).strip().lower() == "doublure":
            df.rename(columns={col: "Doublure"}, inplace=True)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if 'Link' not in header:
        df['link_url'] = None
        return df

    link_col_index = header.index('Link') + 1
    urls = []
    for row in sheet.iter_rows(min_row=2):
        cell = row[link_col_index - 1]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        else:
            match = re.search(r'"(https?://[^"]+)"', str(cell.value))
            urls.append(match.group(1) if match else None)

    df['link_url'] = urls
    return df

def filter_nesten(df, datum_juli_str, datum_sep_str):
    df = df[df['Omschrijving'].str.contains("nest", case=False)].copy()
    df = df[df['Doublure'] != 1.0]
    df['Datum_parsed'] = pd.to_datetime(df['Datum_parsed'], errors='coerce')

    jaar_min = df['Datum_parsed'].dt.year.min()
    datum_juli = datetime.strptime(f"{datum_juli_str}-{jaar_min}", "%d-%m-%Y")
    datum_sep = datetime.strptime(f"{datum_sep_str}-{jaar_min}", "%d-%m-%Y")

    na_sep = df[df['Datum_parsed'] > datum_sep]
    voor_juli = df[df['Datum_parsed'] < datum_juli]
    na_juli = df[df['Datum_parsed'] >= datum_juli]

    unieke_nesten = []
    for _, rij in voor_juli.iterrows():
        coord_i = parse_gps(rij['GPS'])
        if coord_i is None:
            continue
        opvolgers = na_juli['GPS'].apply(lambda x: (
            parse_gps(x) is not None and geodesic(coord_i, parse_gps(x)).meters < 200
        ))
        if not opvolgers.any():
            unieke_nesten.append(rij)

    resultaat = pd.concat([na_sep, pd.DataFrame(unieke_nesten)])
    return resultaat

def maak_kaart(nesten_df, afstand):
    eerste_coord = parse_gps(nesten_df.iloc[0]['GPS'])
    kaart = folium.Map(location=eerste_coord, zoom_start=15)

    for _, rij in nesten_df.iterrows():
        nest_coord = parse_gps(rij['GPS'])
        if nest_coord is None:
            continue

        if str(rij.get('Doublure')).strip().upper() in ["WAAR", "TRUE"]:
            continue

        link = rij.get('link_url') or rij.get('Link')
        datum = rij['Datum_parsed'].strftime("%d-%m-%Y") if pd.notna(rij['Datum_parsed']) else "Onbekend"

        if pd.notna(link) and str(link).startswith("https"):
            popup_html = (
                f"<b>Waarneming ID:</b> {rij['Waarneming ID']}<br>"
                f"<b>Datum:</b> {datum}<br>"
                f"<a href='{link}' target='_blank'>Bekijk melding</a>"
            )
        else:
            popup_html = (
                f"<b>Waarneming ID:</b> {rij['Waarneming ID']}<br>"
                f"<b>Datum:</b> {datum}<br>"
                f"<i>Geen link beschikbaar</i>"
            )

        folium.Marker(
            location=nest_coord,
            popup=folium.Popup(popup_html, max_width=250),
            icon=folium.Icon(color='red', icon='info-sign')
        ).add_to(kaart)

        vallen = genereer_vallen(nest_coord, afstand)
        for val in vallen:
            folium.CircleMarker(
                location=val,
                radius=4,
                color='blue',
                fill=True,
                fill_opacity=0.6
            ).add_to(kaart)

    kaart.save("kaart_nesten_vallen.html")
    webbrowser.open("kaart_nesten_vallen.html")

def selecteer_bestand():
    global bestand_pad
    bestand_pad = filedialog.askopenfilename(filetypes=[("Excel bestanden", "*.xlsx")])
    if bestand_pad:
        messagebox.showinfo("Bestand geselecteerd", f"Je hebt gekozen:\n{bestand_pad}")
    else:
        messagebox.showwarning("Geen bestand", "Selecteer een Excel-bestand.")

def start_gui():
    root = Tk()
    root.title("Nestkaart Generator")

    afstand_var = StringVar(root)
    afstand_var.set("70")
    afstand_opties = [str(x) for x in range(70, 151, 20)]

    Label(root, text="Kies welke afstand tot nest (m) je de vallen wilt plaatsen:").grid(row=0, column=0, sticky="w")
    OptionMenu(root, afstand_var, *afstand_opties).grid(row=0, column=1)

    Label(root, text="Kies datum (vóór 1 juli (dd-mm)) van primaire nesten waarvoor geen secundair nest meer is gevonden:").grid(row=1, column=0, sticky="w")
    datum_juli_entry = Entry(root)
    datum_juli_entry.insert(0, "01-07")
    datum_juli_entry.grid(row=1, column=1)

    Label(root, text="Kies de datum (na 1 september (dd-mm)) van secundaire nesten waar koninginnen zijn uitgevlogen:").grid(row=2, column=0, sticky="w")
    datum_sep_entry = Entry(root)
    datum_sep_entry.insert(0, "01-09")
    datum_sep_entry.grid(row=2, column=1)

    Button(root, text="Selecteer Excel-bestand met waarnemingen", command=selecteer_bestand).grid(row=3, column=0, columnspan=2, pady=5)

    def uitvoeren():
        if not bestand_pad:
            messagebox.showwarning("Geen bestand", "Selecteer eerst een Excel-bestand.")
            return
        try:
            afstand = int(afstand_var.get())
            datum_juli_str = datum_juli_entry.get()
            datum_sep_str = datum_sep_entry.get()

            df = lees_excel_met_links(bestand_pad)
            geselecteerde_nesten = filter_nesten(df, datum_juli_str, datum_sep_str)
            geselecteerde_nesten = geselecteerde_nesten[~geselecteerde_nesten['Doublure'].astype(str).str.strip().str.upper().isin(["WAAR", "TRUE"])]

            if geselecteerde_nesten.empty:
                messagebox.showinfo("Geen nesten", "Er zijn geen geschikte nesten gevonden.")
                return

            maak_kaart(geselecteerde_nesten, afstand)
            messagebox.showinfo("Kaart gereed", "De kaart is opgeslagen en geopend in je browser.")
        except Exception as e:
            messagebox.showerror("Fout", f"Er is een fout opgetreden:\n{e}")

    Button(root, text="Genereer kaart", command=uitvoeren).grid(row=4, column=0, columnspan=2, pady=10)
    root.mainloop()

start_gui()