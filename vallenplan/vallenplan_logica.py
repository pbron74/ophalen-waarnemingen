import math
import os
import re
import webbrowser
import zipfile
from datetime import datetime
from pathlib import Path
from tkinter import (
    Button,
    Entry,
    Label,
    OptionMenu,
    StringVar,
    Tk,
    filedialog,
    messagebox,
)

import folium
import openpyxl
import pandas as pd
import simplekml
from config import DATA_DIR, resource_path
from geopy.distance import geodesic

bestand_pad = None  # Globale variabele voor geselecteerd bestand
ICOON_PAD = resource_path("data/AziatischeHoornaar.jpg")

DEBUG_PAD = os.path.join(DATA_DIR, "debug_nesten.txt")


def debug_log(tekst):
    with open(DEBUG_PAD, "a", encoding="utf-8") as f:
        f.write(str(tekst) + "\n")


def parse_gps(gps_str):
    match = re.search(r"GPS\s*([\d.]+),\s*([\d.]+)", str(gps_str))
    if match:
        return float(match.group(1)), float(match.group(2))
    return None


def genereer_vallen(coord, afstand):
    vallen = []
    for hoek in range(0, 360, 45):
        dx = afstand * math.cos(math.radians(hoek))
        dy = afstand * math.sin(math.radians(hoek))
        lat_offset = dy / 111000
        lon_offset = dx / (111000 * math.cos(math.radians(coord[0])))
        vallen.append((coord[0] + lat_offset, coord[1] + lon_offset))
    return vallen


def lees_excel_met_links(filepath):
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()

    for col in df.columns:
        if str(col).strip().lower() == "doublure":
            df.rename(columns={col: "Doublure"}, inplace=True)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if "Link" not in header:
        df["link_url"] = None
        return df

    link_col_index = header.index("Link") + 1
    urls = []
    for row in sheet.iter_rows(min_row=2):
        cell = row[link_col_index - 1]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        else:
            match = re.search(r'"(https?://[^"]+)"', str(cell.value))
            urls.append(match.group(1) if match else None)

    df["link_url"] = urls

    return df


def filter_nesten(df, datum_voor_str, datum_na_str):
    # 1. Filter op nesten
    df = df[df["Omschrijving"].str.contains("nest", case=False, na=False)].copy()

    # 2. Doublure-filter (Excel-gedrag)
    # Excel toont ONWAAR voor 0.0 en NaN → dus meenemen
    # WAAR = 1.0 → verwijderen
    df = df[df["Doublure"].fillna(0.0) == 0.0]

    # 3. Datumkolom opschonen
    df["Datum"] = (
        df["Datum"]
        .astype(str)
        .str.replace(",0", "", regex=False)
        .str.replace(r"[^\d\-/: ]", "", regex=True)
        .str.strip()
    )

    # 4. Datum converteren
    df["Datum"] = pd.to_datetime(
        df["Datum"], errors="coerce", dayfirst=True, format="mixed"
    )

    # 5. GPS parsen
    df["GPS_parsed"] = df["GPS"].apply(parse_gps)

    # → Alleen rijen met geldige GPS én geldige datum
    df = df[df["GPS_parsed"].notna() & df["Datum"].notna()].copy()

    # 6. Jaar bepalen
    jaar = int(df["Datum"].dt.year.dropna().astype(int).min())

    # 7. GUI-invoer opschonen
    datum_voor_str = re.sub(r"[^0-9\-]", "", datum_voor_str).strip()
    datum_na_str = re.sub(r"[^0-9\-]", "", datum_na_str).strip()

    # 8. Datumgrenzen opbouwen
    datum_voor = datetime.strptime(f"{datum_voor_str}-{jaar}", "%d-%m-%Y")
    datum_na = datetime.strptime(f"{datum_na_str}-{jaar}", "%d-%m-%Y")

    # 9. Selecties
    voor = df[df["Datum"] < datum_voor].reset_index(drop=True)
    na = df[df["Datum"] > datum_na].reset_index(drop=True)

    # 10. Primaire nesten zonder opvolger binnen 200 m
    unieke_primaire = []
    for _, rij in voor.iterrows():
        coord_i = rij["GPS_parsed"]
        opvolgers = na["GPS_parsed"].apply(lambda x: geodesic(coord_i, x).meters < 200)

        if not opvolgers.any():
            unieke_primaire.append(rij)

    # 11. Resultaat samenstellen
    resultaat = pd.concat([na, pd.DataFrame(unieke_primaire)], ignore_index=True)

    # Debug
    debug_log(f"Filter_nesten resultaat: {len(resultaat)} rijen")
    debug_log("UNIEKE DOUBLURE-WAARDEN:")
    debug_log(df["Doublure"].apply(lambda x: (x, type(x))).unique())

    resultaat.to_excel(
        os.path.join(DATA_DIR, "debug_nesten_resultaat.xlsx"), index=False
    )
    debug_log(f"DEBUG: resultaat weggeschreven met {len(resultaat)} rijen")

    return resultaat


def maak_kaart(nesten_df, afstand, excel_pad, datum_juli_str, datum_sep_str):
    debug_log("maak_kaart() START")

    eerste_coord = parse_gps(nesten_df.iloc[0]["GPS"])
    kaart = folium.Map(location=eerste_coord, zoom_start=15)

    jaar_min = int(nesten_df["Datum"].dt.year.dropna().astype(int).min())
    datum_juli = datetime.strptime(f"{datum_juli_str}-{jaar_min}", "%d-%m-%Y")
    datum_sep = datetime.strptime(f"{datum_sep_str}-{jaar_min}", "%d-%m-%Y")

    aantal_nesten = 0
    totaal_vallen = 0

    for _, rij in nesten_df.iterrows():
        # GPS check
        nest_coord = parse_gps(rij["GPS"])
        if nest_coord is None:
            continue

        # Datum check
        datum_val = rij["Datum"]
        if pd.isna(datum_val):
            continue

        # Filter op kleur
        if datum_val < datum_juli:
            kleur = "green"
        elif datum_val > datum_sep:
            kleur = "orange"
        else:
            continue  # tussen juli en september → NIET op kaart

        # → Alleen nesten die OP DE KAART KOMEN tellen
        aantal_nesten += 1

        # Marker tekenen
        link = rij.get("link_url") or rij.get("Link")
        datum_str = datum_val.strftime("%d-%m-%Y")

        popup_html = (
            f"<b>Waarneming ID:</b> {rij['Waarneming ID']}<br>"
            f"<b>Datum:</b> {datum_str}<br>"
            f"{f'<a href={link} target=_blank>Bekijk melding</a>' if pd.notna(link) and str(link).startswith('https') else '<i>Geen link beschikbaar</i>'}"
        )

        folium.Marker(
            location=nest_coord,
            popup=folium.Popup(popup_html, max_width=250),
            icon=folium.Icon(color=kleur, icon="info-sign"),
        ).add_to(kaart)

        # → Vallen genereren (1×) en tellen
        vallen_coords = genereer_vallen(nest_coord, afstand)
        totaal_vallen += len(vallen_coords)

        for val in vallen_coords:
            folium.CircleMarker(
                location=val, radius=4, color="blue", fill=True, fill_opacity=0.6
            ).add_to(kaart)

    # LEGENDA
    legenda_html = f"""
    <div style="position: absolute;
                top: 50px; left: 300px; width: 500px; height: 160px;
                background-color: transparent; border:2px solid grey; z-index:9999; font-size:14px;color: darkred; padding:5px;">
        <b>Legenda</b><br>
        <i style="color:green;">●</i> Primaire nesten (voor {datum_juli_str})<br>
        <i style="color:orange;">●</i> Secundaire nesten (na {datum_sep_str})<br><br>
        <b>Afstand vallen tot nest en onderling:</b> {afstand} meter<br>
        <b>Aantal geselecteerde nesten:</b> {aantal_nesten}<br>
        <b>Totaal aantal vallen:</b> {totaal_vallen}
    </div>
    """
    kaart.get_root().html.add_child(folium.Element(legenda_html))

    # BESTANDSNAAM
    bestandsnaam_input = os.path.basename(excel_pad)
    gemeente_naam = (
        bestandsnaam_input.split("_")[0] if "_" in bestandsnaam_input else "onbekend"
    )
    gemeente_naam = gemeente_naam.replace(" ", "_")
    bestandsnaam_output = f"vallenplan_{gemeente_naam}.html"
    output_path = os.path.join(DATA_DIR, bestandsnaam_output)

    debug_log(f"Nesten op kaart: {aantal_nesten}")
    debug_log(f"Vallen op kaart: {totaal_vallen}")

    debug_log("maak_kaart() EINDE")
    debug_log(f"KAART WORDT OPGESLAGEN ALS: {output_path}")

    kaart.save(output_path)
    webbrowser.open(output_path)
    debug_log(f"KAART GEOPEND: {output_path}")

    return output_path


def maak_kml_per_gemeente(nesten_df, afstand, excel_pad, icoon_pad):
    if "Gemeente" not in nesten_df.columns:
        raise ValueError("Kolom 'Gemeente' ontbreekt in de dataset.")

    icoon_bestandsnaam = os.path.basename(icoon_pad)
    kmz_paden = []

    groepen = nesten_df.groupby("Gemeente")

    for gemeente, groep in groepen:
        kml = simplekml.Kml()
        folder = kml.newfolder(name=str(gemeente))

        for _, rij in groep.iterrows():
            coord = parse_gps(rij["GPS"])
            if coord is None:
                continue

            pnt = folder.newpoint(
                name=f"Nest {rij['Waarneming ID']}", coords=[(coord[1], coord[0])]
            )

            pnt.style.iconstyle.icon.href = icoon_bestandsnaam
            pnt.style.iconstyle.scale = 1.2

            datum_val = rij.get("Datum")
            if pd.notna(datum_val):
                try:
                    datum_str = datum_val.strftime("%d-%m-%Y")
                except Exception:
                    datum_str = str(datum_val)
            else:
                datum_str = "Onbekend"

            pnt.description = (
                f"Waarneming ID: {rij['Waarneming ID']}<br>"
                f"Datum: {datum_str}<br>"
                f"Gemeente: {gemeente}<br>"
            )

            vallen_coords = genereer_vallen(coord, afstand)
            for val in vallen_coords:
                val_pnt = folder.newpoint(name="Val", coords=[(val[1], val[0])])
                val_pnt.style.iconstyle.color = simplekml.Color.blue
                val_pnt.style.iconstyle.scale = 0.7

        # KMZ per gemeente opslaan
        veilige_naam = str(gemeente).replace(" ", "_")
        kmz_naam = f"vallenplan_{veilige_naam}.kmz"
        kmz_pad = os.path.join(DATA_DIR, kmz_naam)

        kml_bytes = kml.kml().encode("utf-8")

        import zipfile

        with zipfile.ZipFile(kmz_pad, "w", zipfile.ZIP_DEFLATED) as kmz:
            kmz.writestr("doc.kml", kml_bytes)
            kmz.write(icoon_pad, icoon_bestandsnaam)

        kmz_paden.append(kmz_pad)

    return kmz_paden


def zip_kmz_bestanden(kmz_paden, uitvoer_map):
    import zipfile
    from datetime import datetime

    os.makedirs(uitvoer_map, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    zip_naam = f"vallenplan_gemeenten_{timestamp}.zip"
    zip_pad = os.path.join(uitvoer_map, zip_naam)

    with zipfile.ZipFile(zip_pad, "w", zipfile.ZIP_DEFLATED) as zipf:
        for kmz in kmz_paden:
            zipf.write(kmz, os.path.basename(kmz))

    return zip_pad


def selecteer_bestand(root, afstand_var, datum_juli_entry, datum_sep_entry):
    global bestand_pad
    bestand_pad = filedialog.askopenfilename(
        initialdir=DATA_DIR, filetypes=[("Excel bestanden", "*.xlsx")]
    )
    if not bestand_pad:
        messagebox.showwarning("Geen bestand", "Selecteer een Excel-bestand.")
        return

    try:
        afstand = int(afstand_var.get())
        df = lees_excel_met_links(bestand_pad)
        geselecteerde_nesten = filter_nesten(
            df, datum_juli_entry.get(), datum_sep_entry.get()
        )

        if geselecteerde_nesten.empty:
            messagebox.showinfo(
                "Geen nesten", "Er zijn geen geschikte nesten gevonden."
            )
        else:
            output_path = maak_kaart(
                geselecteerde_nesten,
                afstand,
                bestand_pad,
                datum_juli_entry.get(),
                datum_sep_entry.get(),
            )

            kmz_paden = maak_kml_per_gemeente(
                geselecteerde_nesten, afstand, bestand_pad, ICOON_PAD
            )
            zip_pad = zip_kmz_bestanden(kmz_paden, DATA_DIR)

            messagebox.showinfo(
                "Kaarten gereed",
                f"De HTML-kaart is opgeslagen in:\n{output_path}\n\n"
                f"ZIP-bestand met alle KMZ’s:\n{zip_pad}\n\n"
                f"De HTML-kaart is geopend in je browser.",
            )

        root.destroy()

    except Exception as e:
        messagebox.showerror("Fout", f"Er is een fout opgetreden:\n{e}")
        root.destroy()


def start_gui():
    open(DEBUG_PAD, "w").close()

    root = Tk()
    root.title("Nestkaart Generator")

    afstand_var = StringVar(root)
    afstand_var.set("70")
    afstand_opties = [str(x) for x in range(70, 151, 20)]

    Label(root, text="Afstand tot nest (m):").grid(row=0, column=0, sticky="w")
    OptionMenu(root, afstand_var, *afstand_opties).grid(row=0, column=1)

    Label(root, text="Datum vóór 1 juli (dd-mm):").grid(row=1, column=0, sticky="w")
    datum_juli_entry = Entry(root)
    datum_juli_entry.insert(0, "01-07")
    datum_juli_entry.grid(row=1, column=1)

    Label(root, text="Datum na 1 september (dd-mm):").grid(row=2, column=0, sticky="w")
    datum_sep_entry = Entry(root)
    datum_sep_entry.insert(0, "01-09")
    datum_sep_entry.grid(row=2, column=1)

    # Bestand selecteren triggert meteen uitvoeren + afsluiten
    Button(
        root,
        text="Selecteer Excel-bestand en genereer kaart",
        command=lambda: selecteer_bestand(
            root, afstand_var, datum_juli_entry, datum_sep_entry
        ),
    ).grid(row=3, column=0, columnspan=2, pady=10)

    root.mainloop()


# Start de GUI als het script direct wordt uitgevoerd
if __name__ == "__main__":
    start_gui()
