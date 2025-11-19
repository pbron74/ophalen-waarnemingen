import pandas as pd
import folium
from geopy.distance import geodesic
import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
import openpyxl
from folium.plugins import FeatureGroupSubGroup
from folium.plugins import GroupedLayerControl

kleuren = [
    'red', 'blue', 'green', 'purple', 'orange', 'darkred', 'lightred',
    'beige', 'darkblue', 'darkgreen', 'cadetblue', 'darkpurple','yellow',
    'pink', 'lightblue', 'lightgreen', 'gray', 'black', 'lightgray'
]

STRAAL_NEST_CLUSTER = 50  # meter

def parse_gps(gps_str):
    match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', str(gps_str))
    if match:
        return float(match.group(1)), float(match.group(2))
    return None, None

def extract_url(cell):
    match = re.search(r'"(https?://[^"]+)"', str(cell))
    return match.group(1) if match else None

def lees_excel_met_links(filepath):
    # Lees basisdata met pandas
    df = pd.read_excel(filepath)

    # Open dezelfde file met openpyxl om hyperlinks te extraheren
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb.active

    # Zoek kolomindex van 'Link'
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if 'Link' not in header:
        df['link_url'] = None
        return df

    link_col_index = header.index('Link') + 1  # openpyxl is 1-based

    urls = []
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell = row[link_col_index - 1]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        else:
            match = re.search(r'"(https?://[^"]+)"', str(cell.value))
            urls.append(match.group(1) if match else None)

    df['link_url'] = urls
    return df
def toon_meldingen(filepath, straal_koppeling):
    df = lees_excel_met_links(filepath)
    df['datum'] = pd.to_datetime(df['Datum_parsed'], errors='coerce')
    # GPS-parsing en afronding
    df[['latitude', 'longitude']] = df['GPS'].map(parse_gps).apply(pd.Series)
    #df['latitude'] = df['latitude'].round(5)
    #df['longitude'] = df['longitude'].round(5)
    df = df.dropna(subset=['latitude', 'longitude'])

    # Detecteer nestkandidaten
    df['is_nest_raw'] = df['Omschrijving'].str.contains('nest', case=False, na=False)
    df['is_nest'] = False
    nestkandidaten = df[df['is_nest_raw']].copy()

    nestkandidaten = df[df['is_nest_raw']].copy()

    for _, groep in nestkandidaten.groupby(['latitude', 'longitude']):
        # Filter op niet-doublure
        niet_doublure = groep[~groep['Doublure'].astype(str).str.strip().str.upper().eq('WAAR')]
        if not niet_doublure.empty:
            df.loc[niet_doublure.index, 'is_nest'] = True
        else:
            df.loc[groep.index, 'is_nest'] = True

    # Fase 1: nestclusters zonder kettingvorming
    nest_df = df[df['is_nest']].copy()
    nest_df['nestcluster'] = pd.NA
    nest_cluster_id = 0

    for i, nest_i in nest_df.iterrows():
        if pd.notna(nest_df.at[i, 'nestcluster']):
            continue
        locatie_i = (nest_i['latitude'], nest_i['longitude'])
        nest_df.at[i, 'nestcluster'] = nest_cluster_id

        for j, nest_j in nest_df.iterrows():
            if i == j or pd.notna(nest_df.at[j, 'nestcluster']):
                continue
            locatie_j = (nest_j['latitude'], nest_j['longitude'])
            afstand = geodesic(locatie_i, locatie_j).meters
            if afstand <= STRAAL_NEST_CLUSTER:
                nest_df.at[j, 'nestcluster'] = nest_cluster_id

        nest_cluster_id += 1

    # Merge nestcluster terug in df
    df = df.merge(nest_df[['latitude', 'longitude', 'nestcluster']], on=['latitude', 'longitude'], how='left')

    # ✅ Robuuste overdracht van is_nest via join
    nest_df = nest_df.set_index(['latitude', 'longitude'])
    df_is_nest = df.set_index(['latitude', 'longitude'])[['is_nest']]
    nest_df = nest_df.drop(columns=['is_nest'], errors='ignore')
    nest_df = nest_df.join(df_is_nest, how='left').reset_index()

    # Fase 2: koppeling aan nestclusters
    df['cluster'] = pd.NA
    cluster_counter = 0

    for nestcluster_id in nest_df['nestcluster'].dropna().unique():
        groep = nest_df[nest_df['nestcluster'] == nestcluster_id]
        laatste_datum = groep['datum'].max()
        nestlocaties = groep[['latitude', 'longitude']].to_numpy()

        for i, row in df.iterrows():
            if pd.isna(row['datum']) or pd.isna(row['latitude']) or pd.isna(row['longitude']):
                continue
            if row['datum'] > laatste_datum + pd.Timedelta(days=14):
                continue
            locatie = (row['latitude'], row['longitude'])
            binnen_radius = any(
                geodesic(locatie, tuple(nest)).meters <= straal_koppeling
                for nest in nestlocaties
            )
            if binnen_radius:
                df.at[i, 'cluster'] = cluster_counter
        cluster_counter += 1

    # Fase 3: clustering zonder nest
    overige_df = df[df['cluster'].isna()].copy()
    overige_df['cluster'] = pd.NA
    verwerkt = set()

    for i, rij_i in overige_df.iterrows():
        if i in verwerkt:
            continue
        locatie_i = (rij_i['latitude'], rij_i['longitude'])
        df.at[i, 'cluster'] = cluster_counter
        verwerkt.add(i)

        for j, rij_j in overige_df.iterrows():
            if j == i or j in verwerkt:
                continue
            locatie_j = (rij_j['latitude'], rij_j['longitude'])
            afstand = geodesic(locatie_i, locatie_j).meters
            if afstand <= straal_koppeling:
                df.at[j, 'cluster'] = cluster_counter
                verwerkt.add(j)

        cluster_counter += 1

    # Fase 4: koppeling van losse meldingen aan dichtstbijzijnde cluster
    losse_meldingen = df[df['cluster'].isna()].copy()
    bestaande_clusters = df[df['cluster'].notna()].copy()

    for i, rij in losse_meldingen.iterrows():
        locatie_i = (rij['latitude'], rij['longitude'])
        min_afstand = float('inf')
        dichtstbij_cluster = None

        for cluster_id in bestaande_clusters['cluster'].dropna().unique():
            groep = bestaande_clusters[bestaande_clusters['cluster'] == cluster_id]
            for _, rij_j in groep.iterrows():
                locatie_j = (rij_j['latitude'], rij_j['longitude'])
                afstand = geodesic(locatie_i, locatie_j).meters
                if afstand < min_afstand:
                    min_afstand = afstand
                    dichtstbij_cluster = cluster_id

        if min_afstand <= 2 * straal_koppeling:
            df.at[i, 'cluster'] = dichtstbij_cluster

    # Kaartopbouw
    m = folium.Map(
        location=[df['latitude'].mean(), df['longitude'].mean()],
        zoom_start=10,
        control_scale=True,
        prefer_canvas=True,
        zoom_control=True
    )
    
    laag_met_nest = folium.FeatureGroup(name="Clusters met nest", overlay=True, control=True)
    laag_zonder_nest = folium.FeatureGroup(name="Clusters zonder nest", overlay=True, control=True)

    cluster_telling = df['cluster'].value_counts()
    clusters_met_nest = set(df[(df['is_nest'] == True) & (df['cluster'].notna())]['cluster'].unique())
    clusters_zonder_nest = set(df['cluster'].dropna().unique()) - clusters_met_nest
    
    unieke_clusters = sorted(df['cluster'].dropna().unique())
    cluster_kleur_map = {
        cid: kleuren[i % len(kleuren)]
        for i, cid in enumerate(unieke_clusters)
    }
    
    # Eerst alle markers toevoegen
    for _, row in df.iterrows():
        cluster_id = row['cluster']
        if pd.isna(cluster_id):
            continue

        kleur = cluster_kleur_map[cluster_id]
        aantal = cluster_telling.loc[cluster_id] if cluster_id in cluster_telling else 0
        locatie = [row['latitude'], row['longitude']]
        link = row.get('link_url') or row.get('Link')
        link_html = f"<br><a href='{link}' target='_blank'>Bekijk waarneming</a>" if pd.notna(link) and str(link).startswith("https") else ""
        popup_text = f"{row['datum'].date()}<br>Cluster: {cluster_id} ({aantal} meldingen)<br>{row['Omschrijving']}{link_html}"

        popup = folium.Popup(popup_text, max_width=300)

        if row['is_nest']:
            marker = folium.Marker(
                location=locatie,
                popup=popup,
                icon=folium.Icon(color=kleur, icon='flag', prefix='fa')
            )
        else:
            marker = folium.Marker(
                location=locatie,
                popup=popup,
                icon=folium.Icon(color=kleur, icon='bug', prefix='fa')
            )

        if cluster_id in clusters_met_nest:
            marker.add_to(laag_met_nest)
        else:
            marker.add_to(laag_zonder_nest)

    # Daarna pas de cirkels toevoegen
    for cluster_id in df['cluster'].dropna().unique():
        groep = df[df['cluster'] == cluster_id]
        if groep.empty:
            continue

        lat_mean = groep['latitude'].mean()
        lon_mean = groep['longitude'].mean()
        kleur = cluster_kleur_map[cluster_id]

        folium.Circle(
            location=[lat_mean, lon_mean],
            radius=straal_koppeling,
            color=kleur,
            fill=False,
            weight=2,
            opacity=1.0,
            options={'interactive': False}
        ).add_to(laag_met_nest if cluster_id in clusters_met_nest else laag_zonder_nest)

    # Voeg subgroepen toe
    laag_met_nest.add_to(m)
    laag_zonder_nest.add_to(m)
    folium.LayerControl(collapsed=False).add_to(m)

    # Periode en straal bovenaan
    start_datum = df['datum'].min().date()
    eind_datum = df['datum'].max().date()
    periode_text = f"Periode: {start_datum} t/m {eind_datum} — koppelingstraal: {straal_koppeling} meter"
    periode_html = f"""
    <div style="
        position: fixed;
        top: 10px;
        left: 20%;
        width: 50%;
        background-color: transparent;
        color: black;
        font-size: 14pt;
        text-align: center;
        padding: 8px;
        border-top: 2px solid gray;
        z-index: 9999;
    ">
        {periode_text}
    </div>
    """

    m.get_root().html.add_child(folium.Element(periode_html))

    # Gemeentenaam uit bestandsnaam
    bestandsnaam_input = os.path.basename(filepath)
    gemeente_naam = bestandsnaam_input.split('_')[0] if '_' in bestandsnaam_input else 'onbekend'
    gemeente_naam = gemeente_naam.replace(' ', '_')
    bestandsnaam_output = f"aziatische_hoornaar_meldingen_{gemeente_naam}.html"
    m.save(bestandsnaam_output)
    print(f"✅ Kaart opgeslagen als {bestandsnaam_output}")

def selecteer_bestand_en_straal():
    def start_verwerking():
        try:
            straal_koppeling = int(straal_var.get())
        except ValueError:
            messagebox.showerror("Fout", "Voer een geldige straal in.")
            return

        filepath = bestand_var.get()
        if not filepath:
            messagebox.showerror("Fout", "Selecteer een Excel-bestand.")
            return

        root.destroy()
        toon_meldingen(filepath, straal_koppeling)

    def kies_bestand():
        pad = filedialog.askopenfilename(
            title="Selecteer Excel-bestand",
            filetypes=[("Excel-bestanden", "*.xlsx *.xls")]
        )
        if pad:
            bestand_var.set(pad)

    root = tk.Tk()
    root.title("Meldingen verwerken")

    # Venster centreren
    venster_breedte = 400
    venster_hoogte = 200
    scherm_breedte = root.winfo_screenwidth()
    scherm_hoogte = root.winfo_screenheight()
    x = (scherm_breedte // 2) - (venster_breedte // 2)
    y = (scherm_hoogte // 2) - (venster_hoogte // 2)
    root.geometry(f"{venster_breedte}x{venster_hoogte}+{x}+{y}")

    # Koppelingstraal
    tk.Label(root, text="Koppelingstraal (meter):").pack(pady=(10, 0))
    straal_var = tk.StringVar(value="600")
    opties = [str(s) for s in range(300, 701, 50)]
    tk.OptionMenu(root, straal_var, *opties).pack()

    # Bestand kiezen
    tk.Label(root, text="Excel-bestand:").pack(pady=(10, 0))
    bestand_var = tk.StringVar()
    tk.Entry(root, textvariable=bestand_var, width=40).pack()
    tk.Button(root, text="Bladeren...", command=kies_bestand).pack(pady=(5, 10))

    # Startknop
    tk.Button(root, text="Verwerken", command=start_verwerking).pack()

    root.mainloop()

selecteer_bestand_en_straal()