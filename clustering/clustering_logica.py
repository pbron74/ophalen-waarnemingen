import pandas as pd
import folium
from geopy.distance import geodesic
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import re
import openpyxl
from folium.plugins import FeatureGroupSubGroup, GroupedLayerControl,MarkerCluster
import os, sys, datetime

def get_data_dir():
    if hasattr(sys, "_MEIPASS"):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(__file__)
    data_dir = os.path.join(base_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir

def log_debug(message):
    data_dir = get_data_dir()
    log_path = os.path.join(data_dir, "debug_log.txt")
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")

def get_base_dir():
    # In een PyInstaller exe bestaat sys._MEIPASS
    if hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(__file__)

# Zorg dat er altijd een data-map is
base_dir = get_base_dir()
data_dir = os.path.join(base_dir, "data")
os.makedirs(data_dir, exist_ok=True)

kleuren = [
    'red', 'blue', 'green', 'purple', 'orange', 'darkred', 'lightred',
    'beige', 'darkblue', 'darkgreen', 'cadetblue', 'darkpurple', 'yellow',
    'pink', 'lightblue', 'lightgreen', 'gray', 'black', 'lightgray'
]

STRAAL_NEST_CLUSTER = 50  # meter

def parse_gps(gps_str):
    match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', str(gps_str))
    if match:
        return float(match.group(1)), float(match.group(2))
    return None, None

def extract_url(cell):
    match = re.search(r'"(https?://[^"]+)"', str(cell))
    return match.group(1) if match else None

def lees_excel_met_links(filepath):
    df = pd.read_excel(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if 'Link' not in header:
        df['link_url'] = None
        return df

    link_col_index = header.index('Link') + 1
    urls = []
    for row in sheet.iter_rows(min_row=2):
        cell = row[link_col_index - 1]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        else:
            match = re.search(r'"(https?://[^"]+)"', str(cell.value))
            urls.append(match.group(1) if match else None)

    df['link_url'] = urls
    return df

def toon_meldingen(filepath, straal_koppeling):
    df = lees_excel_met_links(filepath)
    df['datum'] = pd.to_datetime(df['Datum_parsed'], errors='coerce')
    df[['latitude', 'longitude']] = df['GPS'].map(parse_gps).apply(pd.Series)
    df = df.dropna(subset=['latitude', 'longitude'])
    log_debug(f"Aantal meldingen met GPS: {len(df)}")

    if len(df) >= 2:
        afstand_test = geodesic(
            (df.iloc[0]['latitude'], df.iloc[0]['longitude']),
            (df.iloc[1]['latitude'], df.iloc[1]['longitude'])
        ).meters
        log_debug(f"Afstand berekend tussen eerste twee meldingen: {afstand_test:.1f} meter")

    df['is_nest_raw'] = df['Omschrijving'].str.contains('nest', case=False, na=False)
    df['is_nest'] = False
    nestkandidaten = df[df['is_nest_raw']].copy()

    for _, groep in nestkandidaten.groupby(['latitude', 'longitude']):
        niet_doublure = groep[~groep['Doublure'].astype(str).str.strip().str.upper().eq('WAAR')]
        if not niet_doublure.empty:
            df.loc[niet_doublure.index, 'is_nest'] = True
        else:
            df.loc[groep.index, 'is_nest'] = True

    # Fase 1: nestclusters zonder kettingvorming
    nest_df = df[df['is_nest']].copy()
    nest_df['nestcluster'] = pd.NA
    nest_cluster_id = 0

    for i, nest_i in nest_df.iterrows():
        if pd.notna(nest_df.at[i, 'nestcluster']):
            continue
        locatie_i = (nest_i['latitude'], nest_i['longitude'])
        nest_df.at[i, 'nestcluster'] = nest_cluster_id

        for j, nest_j in nest_df.iterrows():
            if i == j or pd.notna(nest_df.at[j, 'nestcluster']):
                continue
            locatie_j = (nest_j['latitude'], nest_j['longitude'])
            afstand = geodesic(locatie_i, locatie_j).meters
            if afstand <= STRAAL_NEST_CLUSTER:
                nest_df.at[j, 'nestcluster'] = nest_cluster_id

        nest_cluster_id += 1

    df = df.merge(nest_df[['latitude', 'longitude', 'nestcluster']], on=['latitude', 'longitude'], how='left')
    
    # ✅ Robuuste overdracht van is_nest via join
    nest_df = nest_df.set_index(['latitude', 'longitude'])
    df_is_nest = df.set_index(['latitude', 'longitude'])[['is_nest']]
    nest_df = nest_df.drop(columns=['is_nest'], errors='ignore')
    nest_df = nest_df.join(df_is_nest, how='left').reset_index()

    # Fase 2: koppeling aan nestclusters
    df['cluster'] = pd.NA
    cluster_counter = 0

    for nestcluster_id in nest_df['nestcluster'].dropna().unique():
        groep = nest_df[nest_df['nestcluster'] == nestcluster_id]
        laatste_datum = groep['datum'].max()
        nestlocaties = groep[['latitude', 'longitude']].to_numpy()

        for i, row in df.iterrows():
            if pd.isna(row['datum']) or pd.isna(row['latitude']) or pd.isna(row['longitude']):
                continue
            if row['datum'] > laatste_datum + pd.Timedelta(days=14):
                continue
            locatie = (row['latitude'], row['longitude'])
            binnen_radius = any(
                geodesic(locatie, tuple(nest)).meters <= straal_koppeling
                for nest in nestlocaties
            )
            if binnen_radius:
                df.at[i, 'cluster'] = cluster_counter
        cluster_counter += 1

   # Fase 3: clustering zonder nest
    overige_df = df[df['cluster'].isna()].copy()
    overige_df['cluster'] = pd.NA
    verwerkt = set()

    for i, rij_i in overige_df.iterrows():
        if i in verwerkt:
            continue
        locatie_i = (rij_i['latitude'], rij_i['longitude'])
        df.at[i, 'cluster'] = cluster_counter
        verwerkt.add(i)

        for j, rij_j in overige_df.iterrows():
            if j == i or j in verwerkt:
                continue
            locatie_j = (rij_j['latitude'], rij_j['longitude'])
            afstand = geodesic(locatie_i, locatie_j).meters
            if afstand <= straal_koppeling:
                df.at[j, 'cluster'] = cluster_counter
                verwerkt.add(j)

        cluster_counter += 1

    # Fase 4: koppeling van losse meldingen aan dichtstbijzijnde cluste
    losse_meldingen = df[df['cluster'].isna()].copy()
    bestaande_clusters = df[df['cluster'].notna()].copy()

    for i, rij in losse_meldingen.iterrows():
        locatie_i = (rij['latitude'], rij['longitude'])
        min_afstand = float('inf')
        dichtstbij_cluster = None

        for cluster_id in bestaande_clusters['cluster'].dropna().unique():
            groep = bestaande_clusters[bestaande_clusters['cluster'] == cluster_id]
            for _, rij_j in groep.iterrows():
                locatie_j = (rij_j['latitude'], rij_j['longitude'])
                afstand = geodesic(locatie_i, locatie_j).meters
                if afstand < min_afstand:
                    min_afstand = afstand
                    dichtstbij_cluster = cluster_id

        if min_afstand <= 2 * straal_koppeling:
            df.at[i, 'cluster'] = dichtstbij_cluster

    log_debug(f"Meldingen zonder cluster: {df['cluster'].isna().sum()}")
    log_debug(f"Aantal meldingen met cluster: {df['cluster'].notna().sum()}")

    m = folium.Map(
        location=[df['latitude'].mean(), df['longitude'].mean()],
        zoom_start=10,
        control_scale=True,
        prefer_canvas=True,
        zoom_control=True
    )

    laag_met_nest = folium.FeatureGroup(name="Clusters met nest", overlay=True, control=True)
    laag_zonder_nest = folium.FeatureGroup(name="Clusters zonder nest", overlay=True, control=True)

    cluster_telling = df['cluster'].value_counts()
    clusters_met_nest = set(df[(df['is_nest'] == True) & (df['cluster'].notna())]['cluster'].unique())
    clusters_zonder_nest = set(df['cluster'].dropna().unique()) - clusters_met_nest

    unieke_clusters = sorted(df['cluster'].dropna().unique())
    cluster_kleur_map = {
        cid: kleuren[i % len(kleuren)]
        for i, cid in enumerate(unieke_clusters)
    }

    markers_toegevoegd = 0

    for idx, row in df.iterrows():
        try:
            cluster_id = row['cluster']
            if pd.isna(cluster_id):
                continue

            kleur = cluster_kleur_map.get(cluster_id, "blue")
            aantal = cluster_telling.loc[cluster_id] if cluster_id in cluster_telling else 0
            locatie = [row['latitude'], row['longitude']]
            link = row.get('link_url') or row.get('Link')
            link_html = f"<br><a href='{link}' target='_blank'>Bekijk waarneming</a>" if pd.notna(link) and str(link).startswith("https") else ""
            popup_text = f"{row['datum'].date()}<br>Cluster: {cluster_id} ({aantal} meldingen)<br>{row['Omschrijving']}{link_html}"

            popup = folium.Popup(popup_text, max_width=300)
            icon = folium.Icon(color=kleur, icon=('flag' if row['is_nest'] else 'bug'), prefix='fa')

            marker = folium.Marker(location=locatie, popup=popup, icon=icon)

            # ✅ marker toevoegen BINNEN de loop
            if cluster_id in clusters_met_nest:
                marker.add_to(laag_met_nest)
            else:
                marker.add_to(laag_zonder_nest)

            markers_toegevoegd += 1
            
        except Exception as e:
            if 'log_debug' in globals():
                log_debug(f"[ERROR] Marker voor index {idx} niet toegevoegd: {e}")
            else:
                print(f"[ERROR] Marker voor index {idx} niet toegevoegd: {e}")

    # ✅ na afloop van de loop
    log_debug(f"Totaal markers daadwerkelijk toegevoegd: {markers_toegevoegd}")


    # Cirkel per cluster
    for cluster_id in df['cluster'].dropna().unique():
        groep = df[df['cluster'] == cluster_id]
        if groep.empty:
            continue

        lat_mean = groep['latitude'].mean()
        lon_mean = groep['longitude'].mean()
        kleur = cluster_kleur_map[cluster_id]

        folium.Circle(
            location=[lat_mean, lon_mean],
            radius=straal_koppeling,
            color=kleur,
            fill=False,
            weight=2,
            opacity=1.0,
            options={'interactive': False}
        ).add_to(laag_met_nest if cluster_id in clusters_met_nest else laag_zonder_nest)

    # Voeg lagen toe
    laag_met_nest.add_to(m)
    laag_zonder_nest.add_to(m)
    folium.LayerControl(collapsed=False).add_to(m)

    # Periode-overlay
    start_datum = df['datum'].min().date()
    eind_datum = df['datum'].max().date()
    periode_text = f"Periode: {start_datum} t/m {eind_datum} — koppelingstraal: {straal_koppeling} meter"
    periode_html = f"""
    <div style="
        position: fixed;
        top: 10px;
        left: 20%;
        width: 50%;
        background-color: transparent;
        color: black;
        font-size: 14pt;
        text-align: center;
        padding: 8px;
        border-top: 2px solid gray;
        z-index: 9999;
    ">
        {periode_text}
    </div>
    """
    m.get_root().html.add_child(folium.Element(periode_html))

    # Bepaal basisdir (exe of script)
    if hasattr(sys, "_MEIPASS"):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(__file__)

    # Zorg dat er een data-map is
    data_dir = os.path.join(base_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    # Opslaan
    bestandsnaam_input = os.path.basename(filepath)
    gemeente_naam = bestandsnaam_input.split('_')[0] if '_' in bestandsnaam_input else 'onbekend'
    gemeente_naam = gemeente_naam.replace(' ', '_')
    bestandsnaam_output = f"aziatische_hoornaar_meldingen_{gemeente_naam}.html"

    map_path = os.path.join(data_dir, bestandsnaam_output)
    try:
        m.save(map_path)
        if 'log_debug' in globals():
            log_debug(f"✅ Kaart opgeslagen als {map_path}")
        else:
            print(f"✅ Kaart opgeslagen als {map_path}")
    except Exception as e:
        if 'log_debug' in globals():
            log_debug(f"[ERROR] Kaart niet opgeslagen: {e}")
        else:
            print(f"[ERROR] Kaart niet opgeslagen: {e}")

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

def selecteer_bestand_en_straal():
    root = tk.Toplevel()
    root.title("Clustering GUI — LIVE")
    root.geometry("520x300")
    root.minsize(500, 240)

    straal_var = tk.StringVar(value="600")
    bestand_var = tk.StringVar(value="")

    straal_status = tk.StringVar(value="Gekozen straal: 600 m")
    bestand_status = tk.StringVar(value="Nog geen bestand gekozen")
    voortgang_status = tk.StringVar(value="Nog niet gestart")

    def on_straal_change(*_):
        straal_status.set(f"Gekozen straal: {straal_var.get()} m")

    def on_bestand_change(*_):
        pad = bestand_var.get()
        bestand_status.set(f"Gekozen bestand: {pad}" if pad else "Nog geen bestand gekozen")

    straal_var.trace_add("write", on_straal_change)
    bestand_var.trace_add("write", on_bestand_change)

    def kies_bestand():
        pad = filedialog.askopenfilename(
            title="Selecteer Excel-bestand",
            filetypes=[("Excel-bestanden", "*.xlsx *.xls")]
        )
        if pad:
            bestand_var.set(pad)

    def start_verwerking():
        try:
            straal_koppeling = int(straal_var.get())
        except ValueError:
            messagebox.showerror("Fout", "Voer een geldige straal in.")
            return
        filepath = bestand_var.get()
        if not filepath:
            messagebox.showerror("Fout", "Selecteer een Excel-bestand.")
            return

        # voortgang starten
        voortgang_status.set("Verwerken gestart...")
        progress["value"] = 0
        root.update_idletasks()

        # Fase 1: data inlezen
        voortgang_status.set("Stap 1/4: Excel inlezen...")
        progress["value"] = 1
        root.update_idletasks()

        # roep jouw clusteringfunctie aan
        toon_meldingen(filepath, straal_koppeling)

        # Fase 2: clustering
        voortgang_status.set("Stap 2/4: Clustering uitvoeren...")
        progress["value"] = 2
        root.update_idletasks()

        # Fase 3: kaart genereren
        voortgang_status.set("Stap 3/4: Kaart genereren...")
        progress["value"] = 3
        root.update_idletasks()

        # Fase 4: opslaan
        voortgang_status.set("Stap 4/4: Kaart opslaan...")
        progress["value"] = 4
        root.update_idletasks()

        voortgang_status.set("✅ Verwerking afgerond, kaart opgeslagen")
        messagebox.showinfo("Klaar", "De kaart is succesvol opgeslagen.")
        root.destroy()

    pad = {"padx": 10, "pady": 6}

    ttk.Label(root, text="Koppelingstraal (meter):").grid(row=0, column=0, sticky="w", **pad)
    opties = [str(s) for s in range(300, 701, 50)]
    straal_combo = ttk.Combobox(root, textvariable=straal_var, values=opties, state="readonly", width=10)
    straal_combo.grid(row=0, column=1, sticky="w", **pad)
    straal_combo.set(straal_var.get())
    ttk.Label(root, textvariable=straal_status, foreground="blue").grid(row=1, column=0, columnspan=2, sticky="w", **pad)

    ttk.Button(root, text="Kies excel bestand", command=kies_bestand).grid(row=2, column=1, sticky="w", **pad)
    ttk.Label(root, textvariable=bestand_status, foreground="blue").grid(row=3, column=0, columnspan=2, sticky="w", **pad)

    ttk.Button(root, text="Verwerken", command=start_verwerking).grid(row=4, column=0, columnspan=2, pady=14)

    # voortgangsbalk + statuslabel
    progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate", maximum=4)
    progress.grid(row=5, column=0, columnspan=2, pady=10)
    ttk.Label(root, textvariable=voortgang_status, foreground="green").grid(row=6, column=0, columnspan=2, sticky="w", **pad)

    root.columnconfigure(1, weight=1)
    root.grab_set()
    root.focus_force()
    root.lift()
    root.mainloop()