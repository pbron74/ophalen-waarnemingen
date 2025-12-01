from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import os, glob
import re
import time
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from xml.etree.ElementTree import Element, SubElement, tostring, parse as etree_parse
from xml.dom.minidom import parseString
from config import DATA_DIR

kleur_dict = {
    "Individueel": "lightgreen",
    "Nest - niet geruimd": "red",
    "Nest - geruimd": "darkred",
    "koningin": "orange"
}

gemeente_dict = {
    "Vijfheerenlanden": 604514,
    "Utrechtse Heuvelrug": 22885,
    "De Ronde Venen": 22845,
    "Stichtse Vecht": 122099,
    "Utrecht": 22872,
    "Woerden": 22881,
    "Lopik": 22686,
    "De Bilt": 22873,
    "Amersfoort": 22868,
    "Leusden": 22870,
    "Houten": 22685,
    "Zeist": 22886,
    "Wijk bij Duurstede": 22904,
    "Soest": 22867,
    "Rhenen": 22912,
    "Oudewater": 22892,
    "Montfoort": 22888,
    "Bunnik": 22894,
    "Woudenberg": 22879,
    "Baarn": 22849,
    "Eemnes": 22835,
    "Bunschoten": 22842,
    "Nieuwegein": 22902,
    "IJsselstein": 22903,
    "Veenendaal": 675578,
    "Renswoude": 22875
}

def start_browser():
    options = Options()
    # Gebruik platformonafhankelijke venstergrootte
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-default-apps")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--remote-allow-origins=*")

    # macOS-specifieke workaround voor Gatekeeper
    if sys.platform == "darwin":
        os.environ['WDM_LOCAL'] = '1'  # voorkom herhaald downloaden

    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def is_doublure(gps_str, bekende_coords, tolerance=0.0005):
    match = re.search(r'([\d.]+),\s*([\d.]+)', gps_str)
    if not match:
        return False
    lat = float(match.group(1))
    lon = float(match.group(2))
    for known_lat, known_lon in bekende_coords:
        if abs(lat - known_lat) <= tolerance and abs(lon - known_lon) <= tolerance:
            return True
    bekende_coords.add((lat, lon))
    return False

def scrape_en_exporteer(startdatum, einddatum, maandnaam, jaar, gemeente, gemeente_code):
    def herstel_ontbrekende_datum(waarneming):
        origineel = waarneming.get("Datum", "")
        parsed = pd.to_datetime(origineel, errors='coerce')
        if pd.isna(parsed):
            match = re.search(r'(\d{4}-\d{2}-\d{2})', origineel)
            if match:
                datum_hersteld = match.group(1) + " 12:00"
                parsed = pd.to_datetime(datum_hersteld, errors='coerce')
            else:
                parsed = pd.NaT
        return parsed

    waarnemingen = []
    waarneming_ids = set()
    bekende_coords = set()
    driver = start_browser()
    driver_detail = start_browser()
    print("✅ Browsers gestart")
    

    try:
        page = 1
        max_per_sessie = 400
        sessie_teller = 0
        vorige_pagina_ids = set()
        while True:
            target_url = f"https://waarneming.nl/locations/{gemeente_code}/observations/?date_after={startdatum.strftime('%Y-%m-%d')}&date_before={einddatum.strftime('%Y-%m-%d')}&search=Aziatische+hoornaar&sort=date&page={page}"
            print(f"🌐 Bezoek pagina {page}")
            driver.get(target_url)

            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "table.table tbody tr"))
                )
            except:
                print("⚠️ Tabel niet gevonden — stoppen.")
                break

            rows = driver.find_elements(By.CSS_SELECTOR, "table.table tbody tr")
            if not rows:
                print("🚫 Geen waarnemingen op deze pagina — stoppen.")
                break

            pagina_ids = set()
            nieuwe_ids = 0

            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) >= 6:
                    try:
                        link_element = cols[0].find_element(By.TAG_NAME, "a")
                        waarneming_url = link_element.get_attribute("href")
                        match = re.search(r'/observation/(\d+)', waarneming_url)
                        waarneming_id = match.group(1) if match else "Onbekend"
                    except:
                        waarneming_url = ""
                        waarneming_id = "Onbekend"

                    pagina_ids.add(waarneming_id)
                    if waarneming_id in waarneming_ids:
                        continue

                    omschrijving = cols[2].text.strip().lower()
                    is_nest = "nest" in omschrijving
                    in_collectie = "collectie" in omschrijving 
                    is_koningin = "koningin" in omschrijving

                    waarneming_type = (
                        "Nest - geruimd" if is_nest and in_collectie else
                        "Nest - niet geruimd" if is_nest else
                        "koningin" if is_koningin else
                        "Individueel"
                    )

                    gps_raw = ""
                    doublure_status = ""

                    if waarneming_url:
                        try:
                            driver_detail.get(waarneming_url)
                            WebDriverWait(driver_detail, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "span.teramap-coordinates:nth-child(1)"))
                            )
                            gps_element = driver_detail.find_element(By.CSS_SELECTOR, "span.teramap-coordinates:nth-child(1)")
                            gps_raw = gps_element.text.strip()
                            if is_nest:
                                doublure_status = is_doublure(gps_raw, bekende_coords)
                        except Exception as e:
                            print(f"⚠️ GPS niet gevonden voor {waarneming_id}: {e}")
                    
                    try:
                        WebDriverWait(driver_detail, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#observation_status td.validation-status-text"))
                        )
                        validatie_element = driver_detail.find_element(By.CSS_SELECTOR, "#observation_status td.validation-status-text")
                        validatie_status = validatie_element.text.strip()
                        print(f"✅ Validatiestatus voor {waarneming_id}: {validatie_status}")
                    except Exception as e:
                        print(f"⚠️ Validatiestatus niet gevonden voor {waarneming_id}: {e}")
                        validatie_status = ""

                    waarneming = {
                        "Waarneming ID": waarneming_id,
                        "Datum": cols[0].text.strip(),
                        "Omschrijving": omschrijving,
                        "Locatie": cols[3].text.strip(),
                        "Waarnemer": cols[4].text.strip(),
                        "Link": waarneming_url,
                        "GPS": gps_raw,
                        "Doublure": doublure_status,
                        "Type": waarneming_type,
                        "Validatie status": validatie_status
                    }
                    waarneming["Datum_parsed"] = herstel_ontbrekende_datum(waarneming)
                    waarnemingen.append(waarneming)
                    waarneming_ids.add(waarneming_id)
                    nieuwe_ids += 1
                    sessie_teller += 1

            if pagina_ids == vorige_pagina_ids or nieuwe_ids == 0:
                print("📉 Geen nieuwe waarnemingen — stoppen.")
                break
            vorige_pagina_ids = pagina_ids


            if sessie_teller >= max_per_sessie:
                print(f"🔄 {sessie_teller} meldingen bereikt — herstart Chrome-sessies.")
                driver.quit()
                driver_detail.quit()
                time.sleep(2)
                driver = start_browser()
                driver_detail = start_browser()
                print("✅ Nieuwe browsers gestart")
                sessie_teller = 0

            page += 1
            time.sleep(1.5)

    finally:
        driver.quit()
        driver_detail.quit()

    def vind_excelbestand(gemeente):
        patroon = os.path.join(DATA_DIR, f"{gemeente}_aziatische_hoornaar_*.xlsx")
        bestanden = glob.glob(patroon)
        if bestanden:
            bestanden.sort(key=os.path.getmtime, reverse=True)
            return bestanden[0]
        else:
            tijdstempel = datetime.now().strftime("%Y%m%d_%H%M")
            return os.path.join(DATA_DIR, f"{gemeente}_aziatische_hoornaar_{tijdstempel}.xlsx")

    def vind_kmlbestand(gemeente):
        patroon = os.path.join(DATA_DIR, f"{gemeente}_aziatische_hoornaar_*.kml")
        bestanden = glob.glob(patroon)
        if bestanden:
            bestanden.sort(key=os.path.getmtime, reverse=True)
            return bestanden[0]
        else:
            tijdstempel = datetime.now().strftime("%Y%m%d_%H%M")
            return os.path.join(DATA_DIR, f"{gemeente}_aziatische_hoornaar_{tijdstempel}.kml")

    # 📁 Bestanden ophalen
    excel_pad = vind_excelbestand(gemeente)
    kml_pad = vind_kmlbestand(gemeente)

    # 📊 Excel aanvullen zonder overschrijven
    nieuwe_df = pd.DataFrame(waarnemingen)
    nieuwe_df["Link"] = nieuwe_df["Link"].apply(
        lambda url: f'=HYPERLINK("{url}", "Bekijk")' if pd.notna(url) and str(url).strip() else ""
    )

    if not os.path.exists(excel_pad):
        print(f"📄 Excel-bestand niet gevonden — nieuw bestand maken.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Waarnemingen"
        ws.append(list(nieuwe_df.columns))
        wb.save(excel_pad)
        print(f"✅ Excel-bestand opgeslagen op: {excel_pad}")

    wb = load_workbook(excel_pad)
    ws = wb.active

    bestaande_ids = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}
    kolomnamen = list(nieuwe_df.columns)

    for _, row in nieuwe_df.iterrows():
        if row["Waarneming ID"] in bestaande_ids:
            continue
        ws.append([row.get(kolom, "") for kolom in kolomnamen])

    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    doublure_col = next((i for i, cell in enumerate(ws[1], start=1) if cell.value == "Doublure"), None)

    if doublure_col:
        for row in ws.iter_rows(min_row=2):
            cell = row[doublure_col - 1]
            if cell.value is True:
                for c in row:
                    c.fill = fill

    wb.save(excel_pad)
    print(f"📁 Excel aangevuld zonder overschrijven: {os.path.basename(excel_pad)}")

    # 🌍 KML aanvullen
    kleur_dict = {"Nest": "red", "Koningin": "gold", "Waarneming": "yellow"}
    kml_colors = {"red": "ff0000ff", "gold": "ffFFD700", "yellow": "ff00ffff"}
    bestaande_coords = set()

    if not os.path.exists(kml_pad):
        print(f"🌍 KML-bestand niet gevonden — nieuw bestand maken.")
        root = Element('kml', xmlns="http://www.opengis.net/kml/2.2")
        doc = SubElement(root, 'Document')
        kml_str = tostring(root, 'utf-8')
        kml_pretty = parseString(kml_str).toprettyxml(indent="  ")
        with open(kml_pad, "w", encoding="utf-8") as f:
            f.write(kml_pretty)

    try:
        tree = etree_parse(kml_pad)
        root = tree.getroot()
        doc = root.find(".//{http://www.opengis.net/kml/2.2}Document")
    except Exception as e:
        print(f"⚠️ Fout bij inlezen KML-bestand: {e}")
        root = Element('kml', xmlns="http://www.opengis.net/kml/2.2")
        doc = SubElement(root, 'Document')

    for w in waarnemingen:
        gps = w.get("GPS", "")
        match = re.search(r'GPS\s*([\d.]+),\s*([\d.]+)', gps)
        if match:
            lat = float(match.group(1))
            lon = float(match.group(2))
            coord_str = f"{lon},{lat},0"
            desc_str = f"{w['Datum']} - {w['Locatie']}"
            key = (coord_str, desc_str)
            if key in bestaande_coords:
                continue

            placemark = SubElement(doc, 'Placemark')
            SubElement(placemark, 'name').text = w['Type']
            SubElement(placemark, 'description').text = desc_str

            style = SubElement(placemark, 'Style')
            kleur = kleur_dict.get(w["Type"], "yellow")
            icon_style = SubElement(style, 'IconStyle')
            icon = SubElement(icon_style, 'Icon')
            icon_href = SubElement(icon, 'href')
            icon_style_color = SubElement(icon_style, 'color')

            if w["Type"].lower() == "koningin":
                icon_href.text = "https://upload.wikimedia.org/wikipedia/commons/thumb/6/6c/Emoji_u1f451.svg/32px-Emoji_u1f451.svg.png"
                icon_style_color.text = kml_colors.get("gold", "ffFFD700")
            else:
                icon_href.text = "http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png"
                icon_style_color.text = kml_colors.get(kleur, "ffff0000")

            point = SubElement(placemark, 'Point')
            SubElement(point, 'coordinates').text = coord_str

            # ✅ voeg key toe zodat doublures worden onthouden
            bestaande_coords.add(key)


    kml_str = tostring(root, 'utf-8')
    kml_pretty = parseString(kml_str).toprettyxml(indent="  ")
    with open(kml_pad, "w", encoding="utf-8") as f:
        f.write(kml_pretty)

    print(f"🌍 KML aangevuld: {os.path.basename(kml_pad)}")

# ✅ Alleen uitvoeren bij direct runnen
if __name__ == "__main__":
    gemeente = "Utrecht"
    gemeente_code = gemeente_dict[gemeente]
    einddatum = datetime.today()
    startdatum = einddatum - timedelta(weeks=2)
    maandnaam = startdatum.strftime("%B")
    jaar = startdatum.year


    scrape_en_exporteer(startdatum, einddatum, maandnaam, jaar, gemeente, gemeente_code)
